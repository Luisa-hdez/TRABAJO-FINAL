import os
import openpyxl
import math
import random

def create_carpet(semestre,asignatura):
    os.mkdir("TrabajoFinal/Semestre_{}/{}".format(semestre,asignatura[0]))


#funcion para crear el nombre del excel
def create_filename(codigo_curso,nombre,cantidad_estudiantes,codigo_grupo):
    fileName="{}-{}-{}-{}.xlsx".format(codigo_curso,nombre,cantidad_estudiantes,codigo_grupo)
    return fileName

def crear_archivo_excel(nombre_archivo):
    
    workbook = openpyxl.Workbook()

    workbook.save(nombre_archivo)


def crear_archivo_agregar_fila(nombre_archivo,CA,NTE,CC,Notas):
    workbook = openpyxl.Workbook()
    hoja = workbook.active

    print(Notas)

    hoja.append(["Codigo Asignatura CA", "Numero Total de Estudiantes NTE", "Codigo del curso CC", "Nota Asignatura (NA)"])
    hoja.append([CA,NTE,CC,Notas[0]])
    
    hoja_notas = workbook.create_sheet(title="Notas lista")
    hoja_notas.append(["Notas"])

    for nota in Notas[1]:
        print(nota)
        hoja_notas.append([nota])

    workbook.save(nombre_archivo)

def generar_notas(estudiantes):
    aprobados = int(estudiantes*0.7)
    reprobados = int(estudiantes - aprobados)

    notas = []
    for _ in range (aprobados):
        notas.append(round(random.uniform(3.0,5.0),2))
    
    for _ in range(reprobados):
        notas.append(round(random.uniform(0.0,2.9),2))


    promedio = sum(notas) / len(notas)
    return [promedio,notas]
    

def estudiantes_x_salon(asignatura, estudiantes_en_semestre, cupos):
    salones_totales = math.ceil(estudiantes_en_semestre / cupos ) #10


    semestre = asignatura[1]
    create_carpet(semestre,asignatura)


    for salon in range(1,salones_totales+1):
        estudiantesEnSalon = estudiantes_en_semestre / salones_totales
                
        if salon == salones_totales and estudiantesEnSalon % (salones_totales+1) != 0:
            estudiantesEnSalon = estudiantesEnSalon - estudiantesEnSalon % (salones_totales+1)


        generar_notas(estudiantesEnSalon)

        codigo = asignatura[0][0:3] + asignatura[2][0:3] +  "-" + semestre + "-" + asignatura[3] +"-"+ str(salon)

        rutaCarpeta = "TrabajoFinal/Semestre_{}/{}".format(semestre,asignatura[0])
        nombreArchivoExcel = create_filename(codigo,str(asignatura[0]).capitalize().replace(" ",""),math.ceil(estudiantesEnSalon), salon)

        crear_archivo_agregar_fila(rutaCarpeta+"/"+nombreArchivoExcel,codigo,estudiantesEnSalon,salon,generar_notas(estudiantesEnSalon))
